from openpyxl import load_workbook


def load_prizeid2sourceid_dict():
    PRIZEID_COL = 2
    SOURCEID_COL = 4
    dict = {}
    wb = load_workbook(filename='prizeid2sourceid.xlsx')
    ws = wb.active
    for i in range(2, ws.max_row + 1):
        prizeid = ws.cell(row=i, column=PRIZEID_COL).value
        sourceid = ws.cell(row=i, column=SOURCEID_COL).value
        dict[prizeid] = sourceid
    return dict


prizeid_sourceid_dict = load_prizeid2sourceid_dict()